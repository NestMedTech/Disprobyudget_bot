# -*- coding: utf-8 -*-
"""
Telegram Bot for Budget Management (Admin and Accountant Roles)
Dasturlash tili: Python 3
Ma'lumotlar bazasi: SQLite 3
Kutubxonalar: pyTelegramBotAPI, openpyxl, pandas, python-dotenv
"""

import os
import sqlite3
import datetime
import threading
import time
import re
from dotenv import load_dotenv
from telebot import TeleBot, types
import pandas as pd

# Muhit o'zgaruvchilarini yuklash
load_dotenv()

# Bot tokenni shu yerga yozing yoki muhit o'zgaruvchisidan oling
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if TOKEN:
    TOKEN = TOKEN.strip().strip('"').strip("'")

if not TOKEN:
    print("WARNING: TELEGRAM_BOT_TOKEN environment variable is not set. Using fallback token for development.")
    TOKEN = "8770857299:AAGbMQC8TdfDTDCw84tremU9iQ57fI-T54A"

TOKEN = TOKEN.strip()

# Proxy sozlamalari (ixtiyoriy)
PROXY = os.getenv("TELEGRAM_PROXY")
if PROXY:
    PROXY = PROXY.strip().strip('"').strip("'")
    if PROXY:
        from telebot import apihelper
        apihelper.proxy = {'https': PROXY}
        print(f"Proxy faollashtirildi: {PROXY}")

# Custom API URL sozlamalari (ixtiyoriy, O'zbekistondagi serverlar uchun bepul ko'prik)
API_URL = os.getenv("TELEGRAM_API_URL")
if API_URL:
    API_URL = API_URL.strip().strip('"').strip("'")
    if API_URL:
        from telebot import apihelper
        apihelper.API_URL = API_URL
        print(f"Custom API URL faollashtirildi: {API_URL}")

bot = TeleBot(TOKEN)

DB_NAME = "budget_bot.db"

def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

# Ma'lumotlar bazasini yaratish va unga dastlabki ma'lumotlarni yozish
def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if 'users' table schema contains 'bugalter'
    cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='users'")
    users_sql = cursor.fetchone()
    if users_sql and "bugalter" in users_sql[0]:
        print("Migrating users table role check constraint from 'bugalter' to 'buxgalter'...")
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.execute("ALTER TABLE users RENAME TO users_old")
        cursor.execute("""
        CREATE TABLE users (
            telegram_id TEXT PRIMARY KEY,
            username TEXT,
            role TEXT NOT NULL CHECK(role IN ('admin', 'buxgalter'))
        )
        """)
        cursor.execute("""
        INSERT INTO users (telegram_id, username, role)
        SELECT telegram_id, username, CASE WHEN role = 'bugalter' THEN 'buxgalter' ELSE role END FROM users_old
        """)
        cursor.execute("DROP TABLE users_old")
        cursor.execute("PRAGMA foreign_keys=ON")
        conn.commit()
    
    # Foydalanuvchilar jadvali
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        telegram_id TEXT PRIMARY KEY,
        username TEXT,
        role TEXT NOT NULL CHECK(role IN ('admin', 'buxgalter'))
    )
    """)
    
    # Balanslar/Kartalar jadvalini yaratishdan oldin currency ustuni bor-yo'qligini tekshirib migratsiya qilamiz
    cursor.execute("PRAGMA table_info(accounts)")
    columns = [col[1] for col in cursor.fetchall()]
    if columns and "currency" not in columns:
        cursor.execute("ALTER TABLE accounts ADD COLUMN currency TEXT DEFAULT 'UZS'")
        conn.commit()

    # Balanslar/Kartalar jadvali
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        balance INTEGER DEFAULT 0, -- tiyin hisobida (1 so'm = 100 tiyin)
        currency TEXT DEFAULT 'UZS'
    )
    """)
    
    # Kategoriyalar jadvali
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        comment_required INTEGER DEFAULT 0 -- 1: shart, 0: ixtiyoriy
    )
    """)
    
    # Tranzaksiyalar jadvali
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL CHECK(type IN ('kirim', 'chiqim')),
        amount INTEGER NOT NULL, -- tiyin hisobida
        account_id INTEGER NOT NULL,
        category_id INTEGER,
        comment TEXT,
        timestamp TEXT NOT NULL,
        telegram_id TEXT NOT NULL,
        FOREIGN KEY(account_id) REFERENCES accounts(id),
        FOREIGN KEY(category_id) REFERENCES categories(id),
        FOREIGN KEY(telegram_id) REFERENCES users(telegram_id)
    )
    """)
    
    # Admin bormi deb tekshirish va kiritish
    cursor.execute("SELECT * FROM users WHERE role = 'admin'")
    if not cursor.fetchone():
        admin_id = os.getenv("ADMIN_TELEGRAM_ID", "5964436818")
        admin_user = os.getenv("ADMIN_USERNAME", "admin_maqsud")
        cursor.execute("INSERT OR IGNORE INTO users (telegram_id, username, role) VALUES (?, ?, ?)", 
                       (admin_id, admin_user, "admin"))
    
    # Naqd usullarini default kiritish
    cursor.execute("SELECT * FROM accounts WHERE name = 'Naqd'")
    if not cursor.fetchone():
        cursor.execute("INSERT OR IGNORE INTO accounts (name, balance, currency) VALUES (?, ?, ?)", ("Naqd", 0, "UZS"))
        
    cursor.execute("SELECT * FROM accounts WHERE name = 'Naqd (USD)'")
    if not cursor.fetchone():
        cursor.execute("INSERT OR IGNORE INTO accounts (name, balance, currency) VALUES (?, ?, ?)", ("Naqd (USD)", 0, "USD"))

    # Boshlang'ich kategoriyalar
    default_categories = [
        ("Oziq-ovqat", 1),
        ("Transport", 0),
        ("Maosh", 0),
        ("Ijara", 1),
        ("Kommunal", 1)
    ]
    for cat, req in default_categories:
        cursor.execute("INSERT OR IGNORE INTO categories (name, comment_required) VALUES (?, ?)", (cat, req))
        
    conn.commit()
    conn.close()

init_db()

# --- YORDAMCHI FUNKSIYALAR ---

def to_tiyin(amount_str):
    """Foydalanuvchi kiritgan summani tiyinga o'tkazish.
    tiyinh hisobi uchun , ni ham . ham hisoblab undan keyingisini tiyin hisobida hisoblaydi.
    """
    try:
        amount_str = amount_str.replace(" ", "")
        parts = re.split(r'[,.]', amount_str, maxsplit=1)
        if len(parts) == 1:
            val = int(parts[0])
            return val * 100
        else:
            som_val = int(parts[0]) if parts[0] else 0
            tiyin_str = parts[1]
            if not tiyin_str:
                tiyin_val = 0
            else:
                if not tiyin_str.isdigit():
                    return None
                tiyin_val = int(tiyin_str)
            return som_val * 100 + tiyin_val
    except ValueError:
        return None

def from_tiyin(tiyin_amount, currency="UZS"):
    """Tiyinni so'm yoki dollar formatiga o'tkazish.
    Summani e'lon qilishda space separator va valyuta belgisi ishlatiladi.
    """
    som_part = abs(tiyin_amount) // 100
    tiyin_part = abs(tiyin_amount) % 100
    sign = "-" if tiyin_amount < 0 else ""
    
    formatted_som = f"{som_part:,}".replace(",", " ")
    if currency == "USD":
        if tiyin_part > 0:
            return f"{sign}{formatted_som}.{tiyin_part:02d} USD"
        else:
            return f"{sign}{formatted_som} USD"
    else:
        if tiyin_part > 0:
            return f"{sign}{formatted_som}.{tiyin_part:02d} UZS"
        else:
            return f"{sign}{formatted_som} UZS"

def get_user_role(telegram_id):
    conn = get_db_connection()
    user = conn.execute("SELECT role FROM users WHERE telegram_id = ?", (str(telegram_id),)).fetchone()
    conn.close()
    return user["role"] if user else None

def get_text_history(days_or_type, account_id=None):
    """Tanlangan davr va hisob bo'yicha tranzaksiyalar tarixini chiroyli matn shaklida qaytaradi."""
    conn = get_db_connection()
    sql = """
    SELECT t.type, t.amount, t.comment, t.timestamp, a.name as account_name, a.currency as account_currency, c.name as category_name
    FROM transactions t
    LEFT JOIN accounts a ON t.account_id = a.id
    LEFT JOIN categories c ON t.category_id = c.id
    """
    params = []
    
    now = datetime.datetime.now()
    if days_or_type == "daily":
        date_str = now.strftime("%Y-%m-%d")
        sql += " WHERE t.timestamp LIKE ?"
        params.append(f"{date_str}%")
    elif days_or_type == "monthly":
        month_str = now.strftime("%Y-%m")
        sql += " WHERE t.timestamp LIKE ?"
        params.append(f"{month_str}%")
    else:
        try:
            days = int(days_or_type)
            start_date = (now - datetime.timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
            sql += " WHERE t.timestamp >= ?"
            params.append(start_date)
        except ValueError:
            pass
            
    if account_id is not None:
        if "WHERE" in sql:
            sql += " AND t.account_id = ?"
        else:
            sql += " WHERE t.account_id = ?"
        params.append(account_id)
        
    sql += " ORDER BY t.timestamp DESC"
    
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    
    if not rows:
        return "Tranzaksiyalar topilmadi."
        
    text_parts = []
    limit = 30
    for row in rows[:limit]:
        tx_type = row["type"]
        amount = row["amount"]
        comment = row["comment"]
        timestamp_str = row["timestamp"]
        account_name = row["account_name"]
        account_currency = row["account_currency"] or "UZS"
        category_name = row["category_name"]
        
        try:
            dt = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
            formatted_time = dt.strftime("%d.%m.%y %H:%M")
        except Exception:
            formatted_time = timestamp_str
            
        formatted_amount = from_tiyin(amount, account_currency)
        
        is_cash = "naqd" in (account_name or "").lower()
        if is_cash:
            acc_display = account_name
            acc_icon = "💵"
        else:
            digits = "".join(re.findall(r"\d+", account_name or ""))
            if len(digits) >= 4:
                acc_display = f"***{digits[-4:]}"
            else:
                acc_display = account_name
            acc_icon = "💳"
            
        if tx_type == "kirim":
            title = "Popolnenie nalichnymi" if is_cash else "Perevod na kartu"
            memo = f"📝 {comment.upper()}" if comment else "📝 BEEPUL P2P, UZ"
            item = (
                f"🟢 <b>{title}</b>\n"
                f"➕ {formatted_amount}\n"
                f"{acc_icon} {acc_display}\n"
                f"{memo}\n"
                f"🕓 {formatted_time}"
            )
        else:
            title = "Spisanie nalichnyh" if is_cash else "Spisanie c kartu"
            memo = f"📝 {(category_name or 'XARAJAT').upper()}"
            if comment:
                memo += f" - {comment.upper()}"
            else:
                memo += " - UZCARD OTHERS 2 ANY PAYNET, UZ"
            item = (
                f"🔴 <b>{title}</b>\n"
                f"➖ {formatted_amount}\n"
                f"{acc_icon} {acc_display}\n"
                f"{memo}\n"
                f"🕓 {formatted_time}"
            )
        text_parts.append(item)
        
    result_text = "\n\n====================\n\n".join(text_parts)
    if len(rows) > limit:
        result_text += f"\n\n... va yana {len(rows) - limit} ta tranzaksiya mavjud. Barchasini ko'rish uchun Excel hisobotini yuklab oling."
    return result_text

# Foydalanuvchi holatlarini saqlash uchun lug'at (FSM o'rniga oddiy context dict)
user_states = {}

# --- KLAVIATURALAR ---

def get_main_keyboard(role):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    if role == "admin":
        markup.add(
            types.KeyboardButton("📊 Joriy balans"),
            types.KeyboardButton("📜 Xabarnomalar tarixi"),
            types.KeyboardButton("📅 Kunlik hisobot"),
            types.KeyboardButton("➕ Buxgalter qo'shish"),
            types.KeyboardButton("➕ Admin qo'shish")
        )
    elif role == "buxgalter":
        markup.add(
            types.KeyboardButton("📥 Kirim"),
            types.KeyboardButton("📤 Chiqim"),
            types.KeyboardButton("📊 Joriy balans"),
            types.KeyboardButton("⚙️ Balansni sozlash"),
            types.KeyboardButton("📜 Xabarnomalar tarixi"),
            types.KeyboardButton("💳 Kartalar"),
            types.KeyboardButton("📁 Kategoriyalar")
        )
    return markup

# --- BOT INTERACTION START ---

@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    telegram_id = str(message.from_user.id)
    username = message.from_user.username or message.from_user.first_name or "User"
    
    role = get_user_role(telegram_id)
    
    if not role:
        conn = get_db_connection()
        user_count = conn.execute("SELECT COUNT(*) as count FROM users").fetchone()["count"]
        if user_count <= 2:
            conn.execute("INSERT INTO users (telegram_id, username, role) VALUES (?, ?, 'admin')", (telegram_id, username))
            conn.commit()
            role = "admin"
        conn.close()
        
    if role:
        user_states[message.chat.id] = {}
        role_uz = "Admin" if role == "admin" else "Buxgalter"
        escaped_username = username.replace("<", "&lt;").replace(">", "&gt;").replace("&", "&amp;")
        bot.send_message(
            message.chat.id, 
            f"Assalomu alaykum, {escaped_username}! Siz tizimga <b>{role_uz}</b> sifatida kirdingiz.",
            reply_markup=get_main_keyboard(role),
            parse_mode="HTML"
        )
    else:
        bot.send_message(
            message.chat.id,
            "Siz tizimda ro'yxatdan o'tmagansiz. Iltimos, adminga murojaat qiling va Telegram IDingizni bering:\n"
            f"Sizning Telegram ID is: <code>{telegram_id}</code>",
            parse_mode="HTML"
        )

# --- KEYBOARD HANDLERS ---

@bot.message_handler(func=lambda msg: True)
def handle_text_messages(message):
    chat_id = message.chat.id
    telegram_id = str(message.from_user.id)
    role = get_user_role(telegram_id)
    
    if not role:
        bot.send_message(chat_id, "Sizda tizimdan foydalanish huquqi yo'q.")
        return

    text = message.text

    if text in ["❌ Bekor qilish", "🔙 Ortga qaytish"] or text.lower() in ["bekor qilish", "ortga qaytish", "cancel", "back"]:
        user_states[chat_id] = {}
        bot.send_message(chat_id, "Bekor qilindi.", reply_markup=get_main_keyboard(role))
        return

    # --- UMUMIY FUNKSIYALAR (Admin va Buxgalter uchun) ---
    if text == "📊 Joriy balans":
        conn = get_db_connection()
        accounts = conn.execute("SELECT name, balance, currency FROM accounts").fetchall()
        conn.close()
        
        res = "<b>📊 Hozirgi Balanslar Holati:</b>\n\n"
        total_uzs = 0
        total_usd = 0
        for acc in accounts:
            curr = acc['currency'] or 'UZS'
            res += f"💳 <b>{acc['name']}</b>: {from_tiyin(acc['balance'], curr)}\n"
            if curr == "USD":
                total_usd += acc['balance']
            else:
                total_uzs += acc['balance']
            
        res += f"\n💰 <b>JAMI UZS:</b> {from_tiyin(total_uzs, 'UZS')}\n"
        res += f"💰 <b>JAMI USD:</b> {from_tiyin(total_usd, 'USD')}"
        bot.send_message(chat_id, res, parse_mode="HTML")
        return

    elif text == "📜 Xabarnomalar tarixi":
        conn = get_db_connection()
        accounts = conn.execute("SELECT id, name FROM accounts").fetchall()
        conn.close()
        
        if not accounts:
            bot.send_message(chat_id, "Hali hech qanday karta/hisob yaratilmagan.")
            return
            
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton("🗂 Barcha hisoblar (Barchasi birga)", callback_data="hist_all"))
        for acc in accounts:
            markup.add(types.InlineKeyboardButton(f"💳 {acc['name']}", callback_data=f"hist_card_{acc['id']}"))
            
        bot.send_message(
            chat_id,
            "📜 Xabarnomalar tarixi menyusi. Bu yerda CardXabar bot orqali kelgan bildirishnomalar tarixini ko'rishingiz mumkin.\n\nKartani tanlang 🤔",
            reply_markup=markup
        )
        return

    elif text == "📅 Kunlik hisobot":
        if role != "admin":
            bot.send_message(chat_id, "Sizda ushbu amaliyotni bajarish huquqi yo'q.")
            return
        export_and_send_report(chat_id, "daily")
        return
        
    elif text == "📆 Oylik hisobot":
        if role != "admin":
            bot.send_message(chat_id, "Sizda ushbu amaliyotni bajarish huquqi yo'q.")
            return
        export_and_send_report(chat_id, "monthly")
        return

    # --- FAQAT ADMIN FUNKSIYALARI ---
    if role == "admin":
        if text == "➕ Buxgalter qo'shish":
            bot.send_message(chat_id, "Yangi buxgalterning Telegram ID-sini kiriting:")
            user_states[chat_id] = {"action": "wait_accountant_id"}
            return
            
        elif text == "➕ Admin qo'shish":
            bot.send_message(chat_id, "Yangi adminning Telegram ID-sini kiriting:")
            user_states[chat_id] = {"action": "wait_admin_id"}
            return
            
        # Admin text kutish bosqichlari
        elif chat_id in user_states and "action" in user_states[chat_id]:
            action = user_states[chat_id]["action"]
            
            if action == "wait_accountant_id":
                acc_id = text.strip()
                if not acc_id.isdigit():
                    bot.send_message(chat_id, "Xato! Telegram ID faqat raqamlardan iborat bo'lishi kerak. Qaytadan kiriting:")
                    return
                user_states[chat_id] = {"action": "wait_accountant_name", "data": {"telegram_id": acc_id}}
                bot.send_message(chat_id, f"ID {acc_id} uchun buxgalter ismini yoki taxallusini kiriting:")
                return
                
            elif action == "wait_accountant_name":
                acc_name = text.strip()
                data = user_states[chat_id]["data"]
                acc_id = data["telegram_id"]
                
                conn = get_db_connection()
                try:
                    conn.execute(
                        "INSERT INTO users (telegram_id, username, role) VALUES (?, ?, 'buxgalter')",
                        (acc_id, acc_name)
                    )
                    conn.commit()
                    bot.send_message(chat_id, f"Muvaffaqiyatli! Buxgalter {acc_name} (ID: {acc_id}) tizimga qo'shildi.")
                except sqlite3.IntegrityError:
                    bot.send_message(chat_id, "Bu Telegram ID allaqachon ro'yxatdan o'tgan yoki xatolik yuz berdi.")
                finally:
                    conn.close()
                user_states[chat_id] = {}
                return

            elif action == "wait_admin_id":
                adm_id = text.strip()
                if not adm_id.isdigit():
                    bot.send_message(chat_id, "Xato! Telegram ID faqat raqamlardan iborat bo'lishi kerak. Qaytadan kiriting:")
                    return
                user_states[chat_id] = {"action": "wait_admin_name", "data": {"telegram_id": adm_id}}
                bot.send_message(chat_id, f"ID {adm_id} uchun admin ismini yoki taxallusini kiriting:")
                return

            elif action == "wait_admin_name":
                adm_name = text.strip()
                data = user_states[chat_id]["data"]
                adm_id = data["telegram_id"]
                
                conn = get_db_connection()
                try:
                    conn.execute(
                        "INSERT INTO users (telegram_id, username, role) VALUES (?, ?, 'admin')",
                        (adm_id, adm_name)
                    )
                    conn.commit()
                    bot.send_message(chat_id, f"Muvaffaqiyatli! Admin {adm_name} (ID: {adm_id}) tizimga qo'shildi.")
                except sqlite3.IntegrityError:
                    bot.send_message(chat_id, "Bu Telegram ID allaqachon ro'yxatdan o'tgan yoki xatolik yuz berdi.")
                finally:
                    conn.close()
                user_states[chat_id] = {}
                return

    # --- FAQAT BUXGALTER FUNKSIYALARI ---
    if role == "buxgalter":
        if text == "💳 Kartalar":
            conn = get_db_connection()
            accounts = conn.execute("SELECT id, name, balance, currency FROM accounts").fetchall()
            conn.close()
            
            if not accounts:
                card_list_text = "Hozircha hech qanday karta/hisob yaratilmagan."
            else:
                card_list_text = "<b>Mavjud kartalar/hisoblar ro'yxati:</b>\n\n"
                for idx, acc in enumerate(accounts, 1):
                    card_list_text += f"{idx}. <b>{acc['name']}</b> ({from_tiyin(acc['balance'], acc['currency'])})\n"
            
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("➕ Karta qo'shish", callback_data="card_add_new"),
                types.InlineKeyboardButton("❌ Karta o'chirish", callback_data="card_delete_select")
            )
            
            bot.send_message(chat_id, card_list_text, parse_mode="HTML", reply_markup=markup)
            return
            
        elif text in ["⚙️ Balansni sozlash", "⚙️ Byudjetni o'rnatish"]:
            conn = get_db_connection()
            accounts = conn.execute("SELECT id, name, balance, currency FROM accounts").fetchall()
            conn.close()
            
            markup = types.InlineKeyboardMarkup(row_width=2)
            for acc in accounts:
                markup.add(types.InlineKeyboardButton(f"{acc['name']} ({from_tiyin(acc['balance'], acc['currency'])})", callback_data=f"set_budget_acc_{acc['id']}"))
                
            bot.send_message(chat_id, "Balansni sozlamoqchi bo'lgan kartangizni yoki naqd usulini tanlang:", reply_markup=markup)
            return
            
        elif text == "📁 Kategoriyalar":
            conn = get_db_connection()
            categories = conn.execute("SELECT id, name, comment_required FROM categories").fetchall()
            conn.close()
            
            if not categories:
                cat_list_text = "Hozircha hech qanday kategoriya yaratilmagan."
            else:
                cat_list_text = "<b>Mavjud kategoriyalar ro'yxati:</b>\n\n"
                for idx, cat in enumerate(categories, 1):
                    req_text = "izoh majburiy" if cat["comment_required"] == 1 else "izoh ixtiyoriy"
                    cat_list_text += f"{idx}. <b>{cat['name']}</b> ({req_text})\n"
            
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("➕ Kategoriya qo'shish", callback_data="cat_add_new"),
                types.InlineKeyboardButton("❌ Kategoriya o'chirish", callback_data="cat_delete_select")
            )
            
            bot.send_message(chat_id, cat_list_text, parse_mode="HTML", reply_markup=markup)
            return
            
        elif text in ["📥 Kirim", "📤 Chiqim"]:
            tx_type = "kirim" if text == "📥 Kirim" else "chiqim"
            user_states[chat_id] = {"action": "tx_select_account", "data": {"type": tx_type}}
            
            conn = get_db_connection()
            accounts = conn.execute("SELECT id, name, balance, currency FROM accounts").fetchall()
            conn.close()
            
            back_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            back_markup.add(types.KeyboardButton("🔙 Ortga qaytish"))
            bot.send_message(chat_id, "Amaliyot boshlandi. Istalgan vaqtda qaytish uchun pastdagi '🔙 Ortga qaytish' tugmasini bosing:", reply_markup=back_markup)
            
            markup = types.InlineKeyboardMarkup(row_width=2)
            for acc in accounts:
                markup.add(types.InlineKeyboardButton(f"{acc['name']} ({from_tiyin(acc['balance'], acc['currency'])})", callback_data=f"tx_acc_{acc['id']}"))
                
            dir_text = "kirim" if tx_type == "kirim" else "chiqim"
            bot.send_message(chat_id, f"Iltimos, {dir_text} qilinadigan hisobni tanlang:", reply_markup=markup)
            return

        # Buxgalter text kutish bosqichlari
        elif chat_id in user_states and "action" in user_states[chat_id]:
            action = user_states[chat_id]["action"]
            
            if action == "wait_card_name":
                card_name = text.strip()
                if not card_name:
                    bot.send_message(chat_id, "Karta nomi bo'sh bo'lishi mumkin emas. Qayta kiriting:")
                    return
                user_states[chat_id] = {"action": "wait_card_initial_balance", "data": {"name": card_name}}
                bot.send_message(chat_id, f"'{card_name}' kartasi uchun boshlang'ich balansni kiriting (Masalan: 500000 yoki o'tkazib yuborish uchun 0 deb yozing):")
                return
                
            elif action == "wait_card_initial_balance":
                bal_str = text.strip()
                tiyin_val = to_tiyin(bal_str)
                if tiyin_val is None:
                    bot.send_message(chat_id, "Format xato! Iltimos, summani tog'ri kiriting (Masalan: 12500.50):")
                    return
                card_name = user_states[chat_id]["data"]["name"]
                
                conn = get_db_connection()
                try:
                    conn.execute("INSERT INTO accounts (name, balance) VALUES (?, ?)", (card_name, tiyin_val))
                    conn.commit()
                    bot.send_message(chat_id, f"Yangi '{card_name}' kartasi yaratildi! Balans: {from_tiyin(tiyin_val)}.")
                except sqlite3.IntegrityError:
                    bot.send_message(chat_id, f"Bunday nomdagi karta/hisob allaqachon mavjud.")
                finally:
                    conn.close()
                user_states[chat_id] = {}
                return
                
            elif action == "wait_budget_amount":
                acc_id = user_states[chat_id]["data"]["acc_id"]
                acc_name = user_states[chat_id]["data"]["acc_name"]
                bal_str = text.strip()
                tiyin_val = to_tiyin(bal_str)
                if tiyin_val is None:
                    bot.send_message(chat_id, "Format xato! Summani tog'ri kiriting (Masalan: 500000 yoki 500000.00):")
                    return
                
                conn = get_db_connection()
                acc = conn.execute("SELECT currency FROM accounts WHERE id = ?", (acc_id,)).fetchone()
                conn.execute("UPDATE accounts SET balance = ? WHERE id = ?", (tiyin_val, acc_id))
                conn.commit()
                conn.close()
                
                bot.send_message(chat_id, f"Muvaffaqiyatli! '{acc_name}' hisobining balansi {from_tiyin(tiyin_val, acc['currency'])} qilib yangilandi.")
                user_states[chat_id] = {}
                return
                
            elif action == "wait_category_name":
                cat_name = text.strip()
                if not cat_name:
                    bot.send_message(chat_id, "Kategoriya nomi bo'sh bo'lishi mumkin emas. Qayta kiriting:")
                    return
                
                user_states[chat_id] = {"action": "wait_category_comment_required", "data": {"name": cat_name}}
                
                markup = types.InlineKeyboardMarkup(row_width=2)
                markup.add(
                    types.InlineKeyboardButton("Ha, izoh shart", callback_data="cat_req_1"),
                    types.InlineKeyboardButton("Yo'q, ixtiyoriy", callback_data="cat_req_0")
                )
                bot.send_message(chat_id, f"'{cat_name}' kategoriyasiga harajat kiritilganda izoh (izoh yozish) majburiymi?", reply_markup=markup)
                return
                
            elif action == "tx_wait_amount":
                amount_str = text.strip()
                amount_tiyin = to_tiyin(amount_str)
                if amount_tiyin is None or amount_tiyin <= 0:
                    bot.send_message(chat_id, "Mablag'ni noldan katta va to'g'ri formatda kiriting (Masalan: 15000 yoki 15000.00):")
                    return
                
                user_states[chat_id]["data"]["amount"] = amount_tiyin
                tx_type = user_states[chat_id]["data"]["type"]
                
                account_id = user_states[chat_id]["data"]["account_id"]
                conn = get_db_connection()
                acc = conn.execute("SELECT currency FROM accounts WHERE id = ?", (account_id,)).fetchone()
                conn.close()
                
                bot.send_message(chat_id, f"Summa qabul qilindi: {from_tiyin(amount_tiyin, acc['currency'])}")
                
                if tx_type == "kirim":
                    # Kirim uchun kategoriya tanlanmaydi, to'g'ridan-to'g'ri izoh so'raladi
                    user_states[chat_id]["action"] = "tx_wait_comment"
                    bot.send_message(chat_id, "Kirim uchun izoh kiriting (Izoh yozish majburiy):")
                else:
                    # Chiqim uchun kategoriya tanlash
                    user_states[chat_id]["action"] = "tx_select_category"
                    
                    conn = get_db_connection()
                    categories = conn.execute("SELECT id, name FROM categories").fetchall()
                    conn.close()
                    
                    if not categories:
                        bot.send_message(chat_id, "Hali hech qanday kategoriya yaratilmagan. Avval kategoriya yarating!")
                        user_states[chat_id] = {}
                        return
                    
                    markup = types.InlineKeyboardMarkup(row_width=2)
                    for cat in categories:
                        markup.add(types.InlineKeyboardButton(cat["name"], callback_data=f"tx_cat_{cat['id']}"))
                    
                    bot.send_message(chat_id, "Kategoriyani tanlang:", reply_markup=markup)
                return
                
            elif action == "tx_wait_comment":
                comment_text = text.strip()
                save_transaction(chat_id, comment_text, telegram_id)
                return

# --- INLINE CALLBACK HANDLERS ---

@bot.callback_query_handler(func=lambda call: True)
def handle_callback_queries(call):
    chat_id = call.message.chat.id
    data = call.data
    
    # ---------------- KATEGORIYALAR SOZLASH CALLBACKS ----------------
    if data == "cat_add_new":
        user_states[chat_id] = {"action": "wait_category_name"}
        bot.send_message(chat_id, "Yangi kategoriya nomini kiriting (Masalan: 'Kanselyariya'):")
        return

    elif data == "cat_delete_select":
        conn = get_db_connection()
        categories = conn.execute("SELECT id, name FROM categories").fetchall()
        conn.close()
        
        if not categories:
            bot.edit_message_text("O'chirish uchun kategoriyalar mavjud emas.", chat_id, call.message.message_id)
            return
            
        markup = types.InlineKeyboardMarkup(row_width=2)
        for cat in categories:
            markup.add(types.InlineKeyboardButton(f"❌ {cat['name']}", callback_data=f"cat_delete_confirm_{cat['id']}"))
            
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text="O'chirmoqchi bo'lgan kategoriyangizni tanlang:",
            reply_markup=markup
        )
        return

    elif data.startswith("cat_delete_confirm_"):
        cat_id = int(data.split("_")[-1])
        conn = get_db_connection()
        cat = conn.execute("SELECT name FROM categories WHERE id = ?", (cat_id,)).fetchone()
        if cat:
            cat_name = cat["name"]
            conn.execute("DELETE FROM categories WHERE id = ?", (cat_id,))
            conn.commit()
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text=f"✅ Kategoriya '{cat_name}' muvaffaqiyatli o'chirildi!"
            )
        else:
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text="❌ Kategoriya topilmadi."
            )
        conn.close()
        return

    # ---------------- KARTALAR SOZLASH CALLBACKS ----------------
    elif data == "card_add_new":
        user_states[chat_id] = {"action": "wait_card_name"}
        bot.send_message(chat_id, "Yangi karta nomini kiriting (Masalan: 'Hamkorbank'):")
        return

    elif data == "card_delete_select":
        conn = get_db_connection()
        accounts = conn.execute("SELECT id, name, balance, currency FROM accounts").fetchall()
        conn.close()
        
        if not accounts:
            bot.edit_message_text("O'chirish uchun kartalar/hisoblar mavjud emas.", chat_id, call.message.message_id)
            return
            
        markup = types.InlineKeyboardMarkup(row_width=1)
        for acc in accounts:
            markup.add(types.InlineKeyboardButton(f"❌ {acc['name']} ({from_tiyin(acc['balance'], acc['currency'])})", callback_data=f"card_delete_confirm_{acc['id']}"))
            
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text="O'chirmoqchi bo'lgan kartangizni tanlang:",
            reply_markup=markup
        )
        return

    elif data.startswith("card_delete_confirm_"):
        card_id = int(data.split("_")[-1])
        conn = get_db_connection()
        acc = conn.execute("SELECT name FROM accounts WHERE id = ?", (card_id,)).fetchone()
        if acc:
            card_name = acc["name"]
            if card_name in ["Naqd", "Naqd (USD)"]:
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=call.message.message_id,
                    text=f"❌ '{card_name}' hisobini o'chirib bo'lmaydi! Bu tizimning asosiy hisobidir."
                )
                conn.close()
                return
                
            conn.execute("DELETE FROM transactions WHERE account_id = ?", (card_id,))
            conn.execute("DELETE FROM accounts WHERE id = ?", (card_id,))
            conn.commit()
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text=f"✅ '{card_name}' kartasi va unga tegishli barcha tranzaksiyalar muvaffaqiyatli o'chirildi!"
            )
        else:
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text="❌ Karta topilmadi."
            )
        conn.close()
        return

    # Tranzaksiyani bekor qilish
    if data.startswith("undo_tx_"):
        tx_id = int(data.split("_")[-1])
        conn = get_db_connection()
        tx = conn.execute("SELECT * FROM transactions WHERE id = ?", (tx_id,)).fetchone()
        if tx:
            # Revert Account balance
            if tx["type"] == "kirim":
                conn.execute("UPDATE accounts SET balance = balance - ? WHERE id = ?", (tx["amount"], tx["account_id"]))
            else:
                conn.execute("UPDATE accounts SET balance = balance + ? WHERE id = ?", (tx["amount"], tx["account_id"]))
            
            # Delete transaction
            conn.execute("DELETE FROM transactions WHERE id = ?", (tx_id,))
            conn.commit()
            conn.close()
            
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text="✅ Tranzaksiya muvaffaqiyatli bekor qilindi (o'chirildi) va karta balansi qayta tiklandi!"
            )
        else:
            conn.close()
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text="❌ Tranzaksiya allaqachon bekor qilingan yoki topilmadi!"
            )
        return

    # Kategoriya izoh majburiyligi tanlovi
    if data.startswith("cat_req_") and chat_id in user_states and user_states[chat_id].get("action") == "wait_category_comment_required":
        req_val = int(data.split("_")[-1])
        cat_name = user_states[chat_id]["data"]["name"]
        
        conn = get_db_connection()
        try:
            conn.execute("INSERT INTO categories (name, comment_required) VALUES (?, ?)", (cat_name, req_val))
            conn.commit()
            req_text = "Izoh majburiy qilib sozlandi" if req_val == 1 else "Izoh ixtiyoriy qilib sozlandi"
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text=f"Yangi kategoriya yaratildi: '{cat_name}'. {req_text}!"
            )
        except sqlite3.IntegrityError:
            bot.edit_message_text(f"Kategoriya nomi '{cat_name}' allaqachon mavjud.", chat_id, call.message.message_id)
        finally:
            conn.close()
        user_states[chat_id] = {}
        
    # Byudjet o'rnatiladigan kartani tanlash
    elif data.startswith("set_budget_acc_"):
        acc_id = int(data.split("_")[-1])
        conn = get_db_connection()
        acc = conn.execute("SELECT name FROM accounts WHERE id = ?", (acc_id,)).fetchone()
        conn.close()
        
        user_states[chat_id] = {
            "action": "wait_budget_amount",
            "data": {"acc_id": acc_id, "acc_name": acc["name"]}
        }
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text=f"'{acc['name']}' hisobi uchun yangi byudjet summasini kiriting (tiyinlari bilan Masalan: 1250000 yoki 1250000.00):"
        )
        
    # Tranzaksiya hisobini tanlash
    elif data.startswith("tx_acc_") and chat_id in user_states and user_states[chat_id].get("action") == "tx_select_account":
        acc_id = int(data.split("_")[-1])
        conn = get_db_connection()
        acc = conn.execute("SELECT name FROM accounts WHERE id = ?", (acc_id,)).fetchone()
        conn.close()
        
        user_states[chat_id]["data"]["account_id"] = acc_id
        user_states[chat_id]["data"]["account_name"] = acc["name"]
        user_states[chat_id]["action"] = "tx_wait_amount"
        
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text=f"Hisob tanlandi: <b>{acc['name']}</b>.",
            parse_mode="HTML"
        )
        
        bot.send_message(
            chat_id=chat_id,
            text="Endi summani kiriting (Masalan: 15400.00 so'm):"
        )

    # Tranzaksiya kategoriyasini tanlash
    elif data.startswith("tx_cat_") and chat_id in user_states and user_states[chat_id].get("action") == "tx_select_category":
        cat_id = int(data.split("_")[-1])
        conn = get_db_connection()
        cat = conn.execute("SELECT name, comment_required FROM categories WHERE id = ?", (cat_id,)).fetchone()
        conn.close()
        
        user_states[chat_id]["data"]["category_id"] = cat_id
        user_states[chat_id]["data"]["category_name"] = cat["name"]
        user_states[chat_id]["data"]["comment_required"] = cat["comment_required"]
        
        # Izoh berish kerakligini tekshirish
        if cat["comment_required"] == 1:
            user_states[chat_id]["action"] = "tx_wait_comment"
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text=f"Kategoriya: <b>{cat['name']}</b>.\n⚠️ Ushbu kategoriya uchun <b>izoh yozish majburiy</b>. Izohni bu yerga text shaklida yuboring:",
                parse_mode="HTML"
            )
        else:
            # Izoh ixtiyoriy, shuning uchun "Izohni yozmaslik" inline buttonini ham qo'shamiz
            user_states[chat_id]["action"] = "tx_wait_comment"
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("❌ Izoh qoldirmaslik", callback_data="tx_skip_comment"))
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=call.message.message_id,
                text=f"Kategoriya: <b>{cat['name']}</b>.\nKirim/chiqim uchun izoh qoldiring yoki pastdagi tugmani bosing:",
                reply_markup=markup,
                parse_mode="HTML"
            )
            
    # Izoh qoldirishdan voz kechish
    elif data == "tx_skip_comment" and chat_id in user_states and user_states[chat_id].get("action") == "tx_wait_comment":
        save_transaction(chat_id, "", str(call.from_user.id))

    # Barcha hisoblar tarixi uchun davr tanlash
    elif data == "hist_all":
        markup = types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            types.InlineKeyboardButton("1 kun", callback_data="hist_all_period_1"),
            types.InlineKeyboardButton("7 kun", callback_data="hist_all_period_7"),
            types.InlineKeyboardButton("30 kun", callback_data="hist_all_period_30"),
            types.InlineKeyboardButton("45 kun", callback_data="hist_all_period_45"),
            types.InlineKeyboardButton("90 kun", callback_data="hist_all_period_90")
        )
        
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text="📜 Barcha hisoblar bo'yicha tranzaksiyalar tarixi (Karta, Naqd, USD). Davrni tanlang:",
            reply_markup=markup
        )

    # Barcha hisoblar bo'yicha Excel faylini yuborish
    elif data.startswith("hist_all_period_"):
        days = int(data.split("_")[-1])
        
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text="🔄 Barcha hisoblar bo'yicha Excel fayli shakllantirilmoqda, iltimos kuting..."
        )
        
        export_all_history_to_excel(chat_id, days)

    # Xabarnomalar tarixi: kartani tanlashdan keyingi bosqich (davrni tanlash)
    elif data.startswith("hist_card_"):
        account_id = int(data.split("_")[-1])
        
        conn = get_db_connection()
        acc = conn.execute("SELECT name FROM accounts WHERE id = ?", (account_id,)).fetchone()
        conn.close()
        
        card_name = acc["name"] if acc else "Noma'lum"
        
        markup = types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            types.InlineKeyboardButton("1 kun", callback_data=f"hist_period_{account_id}_1"),
            types.InlineKeyboardButton("7 kun", callback_data=f"hist_period_{account_id}_7"),
            types.InlineKeyboardButton("30 kun", callback_data=f"hist_period_{account_id}_30"),
            types.InlineKeyboardButton("45 kun", callback_data=f"hist_period_{account_id}_45"),
            types.InlineKeyboardButton("90 kun", callback_data=f"hist_period_{account_id}_90")
        )
        
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text=f"📜 Xabarnomalar tarixi menyusi ({card_name}). Bu yerda CardXabar bot orqali kelgan bildirishnomalar tarixini ko'rishingiz mumkin.\n\nDavrni tanlang:",
            reply_markup=markup
        )

    # Xabarnomalar tarixi: davrni tanlagandan keyin Excel faylini shakllantirib yuborish
    elif data.startswith("hist_period_"):
        parts = data.split("_")
        account_id = int(parts[2])
        days = int(parts[3])
        
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=call.message.message_id,
            text="🔄 Excel fayli shakllantirilmoqda, iltimos kuting..."
        )
        
        export_history_to_excel(chat_id, account_id, days)

# --- TRANZAKSIYANI SAQLASH VA HABAR BERISH ---

def save_transaction(chat_id, comment_text, telegram_id):
    if chat_id not in user_states:
        return
        
    tx_data = user_states[chat_id]["data"]
    tx_type = tx_data["type"]
    amount = tx_data["amount"]
    account_id = tx_data["account_id"]
    account_name = tx_data["account_name"]
    category_id = tx_data.get("category_id")
    category_name = tx_data.get("category_name")
    comment = comment_text.strip()
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Tranzaksiyani bazaga yozish
    cursor.execute(
        "INSERT INTO transactions (type, amount, account_id, category_id, comment, timestamp, telegram_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (tx_type, amount, account_id, category_id, comment if comment else None, timestamp, telegram_id)
    )
    
    # 2. Hisob joriy balansini yangilash
    if tx_type == "kirim":
        cursor.execute("UPDATE accounts SET balance = balance + ? WHERE id = ?", (amount, account_id))
    else:
        cursor.execute("UPDATE accounts SET balance = balance - ? WHERE id = ?", (amount, account_id))
        
    # 3. Yangi balansni olish
    updated_acc = cursor.execute("SELECT balance, currency FROM accounts WHERE id = ?", (account_id,)).fetchone()
    conn.commit()
    conn.close()
    
    # Notification tahrirlash/yuborish
    formatted_amount = from_tiyin(amount, updated_acc['currency'])
    formatted_time = datetime.datetime.now().strftime("%d.%m.%y %H:%M")
    formatted_balance = from_tiyin(updated_acc['balance'], updated_acc['currency'])
    
    is_cash = "naqd" in account_name.lower()
    if is_cash:
        acc_display = account_name
        acc_icon = "💵"
    else:
        digits = "".join(re.findall(r"\d+", account_name))
        if len(digits) >= 4:
            acc_display = f"***{digits[-4:]}"
        else:
            acc_display = account_name
        acc_icon = "💳"

    if tx_type == "kirim":
        title = "Popolnenie nalichnymi" if is_cash else "Perevod na kartu"
        memo = f"📝 {comment.upper()}" if comment else "📝 BEEPUL P2P, UZ"
        msg_notify = (
            f"🟢 <b>{title}</b>\n"
            f"➕ {formatted_amount}\n"
            f"{acc_icon} {acc_display}\n"
            f"{memo}\n"
            f"🕓 {formatted_time}\n"
            f"💵 {formatted_balance}"
        )
    else:
        title = "Spisanie nalichnyh" if is_cash else "Spisanie c kartu"
        memo = f"📝 {category_name.upper()}"
        if comment:
            memo += f" - {comment.upper()}"
        else:
            memo += " - UZCARD OTHERS 2 ANY PAYNET, UZ"
            
        msg_notify = (
            f"🔴 <b>{title}</b>\n"
            f"➖ {formatted_amount}\n"
            f"{acc_icon} {acc_display}\n"
            f"{memo}\n"
            f"🕓 {formatted_time}\n"
            f"💵 {formatted_balance}"
        )
    
    # Buxgalter va barcha adminsga habar berish (klaviaturani qaytarish bilan)
    role = get_user_role(telegram_id)
    bot.send_message(chat_id, msg_notify, parse_mode="HTML", reply_markup=get_main_keyboard(role))
    
    # Boshqa adminlarga ham habar berib turish tizimi
    notify_admins_about_tx(msg_notify, exclude_chat_id=chat_id)
    
    # Holatni tozalash
    user_states[chat_id] = {}

def notify_admins_about_tx(message_text, exclude_chat_id):
    conn = get_db_connection()
    admins = conn.execute("SELECT telegram_id FROM users WHERE role = 'admin'").fetchall()
    conn.close()
    
    for admin in admins:
        admin_id = int(admin["telegram_id"])
        if admin_id != exclude_chat_id:
            try:
                bot.send_message(admin_id, message_text, parse_mode="HTML")
            except Exception:
                pass # Bot o'sha admin bilan avval suhbatlashmagan bo'lishi mumkin

# --- EXCEL EXPORT VA YUBORISH ---

def format_excel_file(file_path):
    """Excel faylini qizil shapka, oq qalin yozuvlar va chiroyli dizayn bilan formatlaydi."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Gridlines (katakchalar chiziqlari) ko'rinib turishi uchun
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
        
        # Qizil fon va oq qalin shrift (shablonga mos holda)
        header_fill = PatternFill(start_color="E30613", end_color="E30613", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # Sarlavhani formatlash
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # Ma'lumot qatorlarini formatlash
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=11)
                
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name == "Summa":
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
                elif col_name in ["Sana", "Vaqt", "Kirim/Chiqim"]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
        # Ustunlar kengligini avtomatik moslash
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = ""
                if cell.value is not None:
                    if isinstance(cell.value, float):
                        val_str = f"{cell.value:,.2f}"
                    else:
                        val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(file_path)
    except Exception as e:
        print(f"Excel faylini bezashda xatolik: {e}")

def export_history_to_excel(chat_id, account_id, days):
    conn = get_db_connection()
    
    # Karta nomini olish
    acc = conn.execute("SELECT name FROM accounts WHERE id = ?", (account_id,)).fetchone()
    card_name = acc["name"] if acc else "Karta"
    
    # Boshlang'ich sanani hisoblash
    now = datetime.datetime.now()
    start_date = (now - datetime.timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    
    sql_query = """
    SELECT 
        a.name as [To'lov turi],
        strftime('%d.%m.%Y', t.timestamp) as [Sana],
        strftime('%H:%M', t.timestamp) as [Vaqt],
        (CAST(t.amount as REAL) / 100.0) as [Summa],
        a.currency as [Valyuta],
        CASE t.type WHEN 'kirim' THEN 'Kirim' ELSE 'Chiqim' END as [Kirim/Chiqim],
        COALESCE(t.comment, 'Izohsiz') as [Nima uchun],
        COALESCE(c.name, 'Kategoriyasiz') as [Kategoriya]
    FROM transactions t
    LEFT JOIN accounts a ON t.account_id = a.id
    LEFT JOIN categories c ON t.category_id = c.id
    WHERE t.account_id = ? AND t.timestamp >= ?
    ORDER BY t.timestamp DESC
    """
    
    df = pd.read_sql_query(sql_query, conn, params=[account_id, start_date])
    conn.close()
    
    if df.empty:
        bot.send_message(chat_id, f"❌ '{card_name}' bo'yicha oxirgi {days} kunlik tranzaksiyalar topilmadi.")
        return
        
    safe_card_name = re.sub(r'[\\/*?:"<>|]', '_', card_name)
    safe_card_name = re.sub(r'_+', '_', safe_card_name).replace(' ', '_')
    file_name = f"Tarix_{safe_card_name}_{days}_kun.xlsx"
    excel_path = os.path.join(".", file_name)
    df.to_excel(excel_path, index=False, sheet_name="Moliya tarixi")
    
    # Chiroyli formatlashni qo'llash
    format_excel_file(excel_path)
    
    with open(excel_path, 'rb') as doc:
        bot.send_document(
            chat_id, 
            doc, 
            caption=f"📊 <b>{card_name}</b> bo'yicha oxirgi <b>{days} kunlik</b> tranzaksiyalar tarixi (Excel formatida).",
            parse_mode="HTML"
        )
        
    try:
        os.remove(excel_path)
    except Exception:
        pass

def export_all_history_to_excel(chat_id, days):
    conn = get_db_connection()
    
    # Boshlang'ich sanani hisoblash
    now = datetime.datetime.now()
    start_date = (now - datetime.timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    
    sql_query = """
    SELECT 
        a.name as [To'lov turi],
        strftime('%d.%m.%Y', t.timestamp) as [Sana],
        strftime('%H:%M', t.timestamp) as [Vaqt],
        (CAST(t.amount as REAL) / 100.0) as [Summa],
        a.currency as [Valyuta],
        CASE t.type WHEN 'kirim' THEN 'Kirim' ELSE 'Chiqim' END as [Kirim/Chiqim],
        COALESCE(t.comment, 'Izohsiz') as [Nima uchun],
        COALESCE(c.name, 'Kategoriyasiz') as [Kategoriya]
    FROM transactions t
    LEFT JOIN accounts a ON t.account_id = a.id
    LEFT JOIN categories c ON t.category_id = c.id
    WHERE t.timestamp >= ?
    ORDER BY t.timestamp DESC
    """
    
    df = pd.read_sql_query(sql_query, conn, params=[start_date])
    conn.close()
    
    if df.empty:
        bot.send_message(chat_id, f"❌ Oxirgi {days} kunlik tranzaksiyalar topilmadi.")
        return
        
    file_name = f"Tarix_Barcha_Hisoblar_{days}_kun.xlsx"
    excel_path = os.path.join(".", file_name)
    df.to_excel(excel_path, index=False, sheet_name="Moliya tarixi")
    
    # Chiroyli formatlashni qo'llash
    format_excel_file(excel_path)
    
    with open(excel_path, 'rb') as doc:
        bot.send_document(
            chat_id, 
            doc, 
            caption=f"📊 <b>Barcha hisoblar</b> bo'yicha oxirgi <b>{days} kunlik</b> tranzaksiyalar tarixi (Excel formatida).",
            parse_mode="HTML"
        )
        
    try:
        os.remove(excel_path)
    except Exception:
        pass

def export_and_send_report(chat_id, report_type):
    conn = get_db_connection()
    
    sql_query = """
    SELECT 
        a.name as [To'lov turi],
        strftime('%d.%m.%Y', t.timestamp) as [Sana],
        strftime('%H:%M', t.timestamp) as [Vaqt],
        (CAST(t.amount as REAL) / 100.0) as [Summa],
        a.currency as [Valyuta],
        CASE t.type WHEN 'kirim' THEN 'Kirim' ELSE 'Chiqim' END as [Kirim/Chiqim],
        COALESCE(t.comment, 'Izohsiz') as [Nima uchun],
        COALESCE(c.name, 'Kategoriyasiz') as [Kategoriya]
    FROM transactions t
    LEFT JOIN accounts a ON t.account_id = a.id
    LEFT JOIN categories c ON t.category_id = c.id
    """
    
    now = datetime.datetime.now()
    if report_type == "daily":
        date_str = now.strftime("%Y-%m-%d")
        sql_query += f" WHERE t.timestamp LIKE '{date_str}%'"
        file_name = f"Kunlik_hisobot_{date_str}.xlsx"
        subject = f"📅 Bugungi kunlik hisobot ({date_str}):"
    else:
        month_str = now.strftime("%Y-%m")
        sql_query += f" WHERE t.timestamp LIKE '{month_str}%'"
        file_name = f"Oylik_hisobot_{month_str}.xlsx"
        subject = f"📆 Shu oylik hisobot ({now.strftime('%B %Y')}):"
        
    sql_query += " ORDER BY t.timestamp DESC"
    
    df = pd.read_sql_query(sql_query, conn)
    conn.close()
    
    if df.empty:
        bot.send_message(chat_id, f"❌ Tanlangan {report_type == 'daily' and 'kun' or 'oy'} uchun hisobot ma'lumotlari topilmadi.")
        return
        
    excel_path = os.path.join(process_cwd_fallback(), file_name)
    df.to_excel(excel_path, index=False, sheet_name="Moliya hisoboti")
    
    # Chiroyli formatlashni qo'llash
    format_excel_file(excel_path)
    
    with open(excel_path, 'rb') as doc:
        bot.send_document(
            chat_id, 
            doc, 
            caption=f"{subject}\nJami tranzaksiyalar soni: {len(df)} ta.",
            reply_to_message_id=None
        )
        
    try:
        os.remove(excel_path)
    except Exception:
        pass

    # Integratsiya: Xabarnomalar tarixini matn shaklida ham jo'natish
    history_text = get_text_history(report_type)
    bot.send_message(
        chat_id,
        f"<b>📋 Xabarnomalar tarixi ({report_type == 'daily' and 'Kunlik' or 'Oylik'}):</b>\n\n{history_text}",
        parse_mode="HTML"
    )

def process_cwd_fallback():
    return "."

def auto_monthly_scheduler():
    """Background scheduler thread to check if today is the 30th of the month at 9:00 AM"""
    print("Background scheduler for monthly auto-reports started.")
    last_sent_month = None
    while True:
        now = datetime.datetime.now()
        if now.day == 30 and now.hour == 9 and last_sent_month != now.month:
            last_sent_month = now.month
            print("Automatic scheduler triggered. Sending monthly reports to admins and accountants...")
            conn = get_db_connection()
            staff = conn.execute("SELECT telegram_id FROM users WHERE role IN ('admin', 'buxgalter')").fetchall()
            conn.close()
            for user in staff:
                try:
                    export_and_send_report(int(user["telegram_id"]), "monthly")
                except Exception as e:
                    print(f"Error sending monthly report to {user['telegram_id']}: {e}")
        time.sleep(600)

from http.server import BaseHTTPRequestHandler, HTTPServer

class DummyServer(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-type", "text/plain")
        self.end_headers()
        self.wfile.write(b"Bot is running!")
    def log_message(self, format, *args):
        pass

def run_dummy_server():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), DummyServer)
    print(f"Dummy port server started on port {port}")
    server.serve_forever()

if __name__ == "__main__":
    scheduler_thread = threading.Thread(target=auto_monthly_scheduler, daemon=True)
    scheduler_thread.start()
    
    # Start dummy HTTP server for Render port binding
    threading.Thread(target=run_dummy_server, daemon=True).start()
    
    print("Bot muvaffaqiyatli ishga tushdi! Telegram botni kuzatishingiz mumkin.")
    bot.infinity_polling()
